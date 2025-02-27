import pandas as pd
import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import tkinter as tk
from tkinter import filedialog
import dashboard as db
import time

log_folder = "Logs"
os.makedirs(log_folder, exist_ok=True)
# Set up logging
logging.basicConfig(
    level=logging.DEBUG, 
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('Logs/Comparison.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger()

# Declare global variables
file_entry_previous = None
file_entry_latest = None

calculated_amount = []
# calculated_totals = 0

def load_sheets(file_previous, file_latest):
    try:
        logger.info("Loading sheets from the provided Excel files.")

        # Define possible sheet names
        sheet_options = {
            # "PR": ["PR"],
            "VDPMV": ["SumVDPMVReport", "SumUberReport"],
            "Trapeze": ["TrapezeReport"],
            # "Partner": ["Div2PartnerList"],
            # "Deductions": ["Deduction and other Revenue"]
        }
        
        def load_sheet(file, sheet_names):
            """Tries to load the first available sheet from a list of sheet names."""
            available_sheets = pd.ExcelFile(file).sheet_names
            for sheet in sheet_names:
                if sheet in available_sheets:
                    return pd.read_excel(file, sheet_name=sheet)
            logger.warning(f"None of the specified sheets {sheet_names} found in {file}")
            return None  # Return None if no matching sheet is found
        
        sheet_pr_previous = pd.read_excel(file_previous, sheet_name="PR")
        sheet_pr_latest = pd.read_excel(file_latest, sheet_name="PR")

        sheet_trapeze_previous = load_sheet(file_previous, sheet_options["Trapeze"])
        sheet_trapeze_latest = load_sheet(file_latest, sheet_options["Trapeze"])

        sheet_vdpmv_previous = load_sheet(file_previous, sheet_options["VDPMV"])
        sheet_vdpmv_latest = load_sheet(file_latest, sheet_options["VDPMV"])
        
        # sheet_vdpmv_previous = pd.read_excel(file_previous, sheet_name="SumVDPMVReport")
        # sheet_vdpmv_latest = pd.read_excel(file_latest, sheet_name="SumVDPMVReport")

        # sheet_partner_previous = pd.read_excel(file_previous, sheet_name="Div2PartnerList")
        # sheet_partner_latest = pd.read_excel(file_latest, sheet_name="Div2PartnerList")

        sheet_deductions_previous = pd.read_excel(file_previous, sheet_name="Deductions")
        sheet_deductions_latest = pd.read_excel(file_latest, sheet_name="Deductions")
        
        logger.info("Sheets loaded successfully.")
        return sheet_pr_previous, sheet_pr_latest, sheet_vdpmv_previous, sheet_vdpmv_latest, sheet_trapeze_previous, sheet_trapeze_latest, sheet_deductions_previous, sheet_deductions_latest
    except Exception as e:
        logger.error(f"Error loading sheets: {e}")
        raise

def clean_currency(value):
    try:
        if isinstance(value, str):
            value = value.replace('$', '').replace(',', '').strip()
            return round(float(value), 2) if value else None
        return value
    except ValueError:
        logger.error(f"Error cleaning currency value: {value}")
        return None

def calculate_totals(deductions_sheet, pr_sheet):
    try:
        calculated_totals = 0
        clients = deductions_sheet["Type"].unique()
        print("Clients in deductions_sheet:", clients)
        for client in clients: 
            # logger.info(f"Calculating totals for {client} clients.")

            if client == "TRANSDEV":
                t_client = "DIV10_TRIMET"
            elif client == "UBER":
                t_client = "UBER WAV TRANSIT"

            client_header_row = pr_sheet[pr_sheet.iloc[:, 0].astype(str).str.contains(t_client, na=False, case=False)]
            if client_header_row.empty:
                logger.warning(f"No header row found for client {t_client} in pr_sheet.")
                continue  # Skip to the next client
            # Get the index of the client header row
            client_header_index = client_header_row.index[0]
            # logger.info(f"Found client header for {client} at row {client_header_index}.")

            # Find the first empty row after the client header row to determine the endpoint
            empty_row_index = pr_sheet.iloc[client_header_index + 1:, 0].isna().idxmax()

            if pd.isna(empty_row_index):
                # If no empty row is found, use the last row of the DataFrame
                next_header_index = pr_sheet.shape[0]
            else:
                # The index of the first empty row marks the endpoint
                next_header_index = empty_row_index + client_header_index + 1

            # logger.info(f"Partner rows for client {client} are between rows {client_header_index + 1} and {next_header_index - 1}.")

            # Extract the partner rows between the client header and the first empty row
            partner_rows = pr_sheet.iloc[client_header_index + 1:next_header_index]

            # Get all unique partners for this client from the deductions sheet
            total_partners = deductions_sheet["PARTNER"].unique()
            # logger.info(f"Total partners for {client}: {len(total_partners)}")

            total_amount = 0

            # Calculate total amount for matching partners
            for partner in total_partners:
                # Find rows in partner_rows matching the partner
                partner_rows_matched = partner_rows[partner_rows.iloc[:, 0].astype(str).str.strip() == str(partner).strip()]

                if not partner_rows_matched.empty:
                    # Add the amount found in column 13 (assuming it's the amount column)
                    total_amount += partner_rows_matched.iloc[0, 13]  # Column 13 contains the amount
                    # logger.info(f"Amount for partner {partner}: {partner_rows_matched.iloc[0, 14]}")
            calculated_totals += total_amount
        logger.info(f"Total amount for client {client}: {total_amount}")
        # logger.info("Calculating totals for trips, hours, operators, and amounts.")
        logger.info(f"Total calculated mamount:{calculated_totals}")
        return calculated_totals
    except Exception as e:
        logger.error(f"Error calculating totals: {e}")
        raise

def calculate_client_totals(deductions_sheet, pr_sheet, client):
    global calculated_amount
    try:
        calculated_totals = 0
        # clients = deductions_sheet["Type"].unique()
        # print("Clients in deductions_sheet:", clients)
        # for client in clients: 
        # logger.info(f"Calculating totals for {client} clients.")

        client_header_row = pr_sheet[pr_sheet.iloc[:, 0].astype(str).str.contains(client, na=False, case=False)]
        if client_header_row.empty:
            logger.warning(f"No header row found for client {client} in pr_sheet.")
            return calculated_totals; 
    
        # Get the index of the client header row
        client_header_index = client_header_row.index[0]
        logger.info(f"Found client header for {client} at row {client_header_index}.")

        # Find the first empty row after the client header row to determine the endpoint
        empty_row_index = pr_sheet.iloc[client_header_index + 1:, 0].isna().idxmax()

        if pd.isna(empty_row_index):
            # If no empty row is found, use the last row of the DataFrame
            next_header_index = pr_sheet.shape[0]
        else:
            # The index of the first empty row marks the endpoint
            next_header_index = empty_row_index + client_header_index + 1

        logger.info(f"Partner rows for client {client} are between rows {client_header_index + 1} and {next_header_index - 1}.")

        # Extract the partner rows between the client header and the first empty row
        partner_rows = pr_sheet.iloc[client_header_index + 1:next_header_index]

        # Get all unique partners for this client from the deductions sheet
        total_partners = deductions_sheet["PARTNER"].unique()
        logger.info(f"Total partners for {client}: {len(total_partners)}")

        total_amount = 0

        # Calculate total amount for matching partners
        for partner in total_partners:
            # Find rows in partner_rows matching the partner
            partner_rows_matched = partner_rows[partner_rows.iloc[:, 0].astype(str).str.strip() == str(partner).strip()]
            # logger.info(f"Partner rows matched: {partner_rows_matched}")
            logger.info(f"Partner: {partner}")

            if not partner_rows_matched.empty:
                # Add the amount found in column 14 (assuming it's the amount column)
                total_amount += partner_rows_matched.iloc[0, 13]  # Column 16 contains the amount
                logger.info(f"Amount for partner {partner}: {partner_rows_matched.iloc[0, 13]}")
                logger.info(f"Current Amount for partner {partner}: {total_amount}")
        calculated_totals += total_amount


        logger.info(f"Total amount for client {client}: {total_amount}")
        # logger.info("Calculating totals for trips, hours, operators, and amounts.")
        logger.info(f"Total calculated mamount:{calculated_totals}")
        return calculated_totals
    except Exception as e:
        logger.error(f"Error calculating cllient totals: {e}")
        raise

def compare_totals(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing Totals between previous and latest values.")

        # Ensure inputs are numeric
        if not isinstance(sheet_previous, (int, float)) or not isinstance(sheet_latest, (int, float)):
            raise TypeError("Both sheet_previous and sheet_latest must be numeric values.")
        
        # Create a DataFrame to store results
        deductions_comparison = pd.DataFrame({
            "LATEST": [sheet_latest],
            "PREVIOUS": [sheet_previous],
            "DIFFERENCE": [sheet_latest - sheet_previous]
        }).round(2)

        # Add CHANGE column based on the difference
        deductions_comparison["CHANGE"] = deductions_comparison["DIFFERENCE"].apply(
            lambda diff: "Increased" if diff > 0 else "Decreased" if diff < 0 else "No Change"
        )

        logger.info("Totals comparison completed.")
        return deductions_comparison

    except Exception as e:
        logger.error(f"Error comparing totals: {e}")
        raise

def compare_client_htotalrev(sheet_previous, sheet_latest, client):
    try:
        logger.info("Comparing Hourly Total Revs between previous and latest sheets.")
        logger.info(f"HTOTAL Client: {client}")
        # Group by "PARTNER NAME" and sum "Total Rev"
        revtable = None
        previous_grouped = None
        latest_grouped = None
        if client == "TRANSDEV":
            revtable = "TTL Rev"
            previous_grouped = sheet_previous.groupby("PARTNER NAME", as_index=False)[revtable].sum()
            latest_grouped = sheet_latest.groupby("PARTNER NAME", as_index=False)[revtable].sum()
        else:
            revtable = "Total Rev"
            previous_grouped = sheet_previous.groupby("PARTNER NAME", as_index=False)[revtable].sum()
            latest_grouped = sheet_latest.groupby("PARTNER NAME", as_index=False)[revtable].sum()

        # Merge both grouped dataframes on "PARTNER NAME"
        comparison = latest_grouped.merge(previous_grouped, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")).fillna(0)

        # Calculate the change in "Total Rev"
        comparison["CHANGE"] = comparison[f"{revtable}_LATEST"] - comparison[f"{revtable}_PREVIOUS"]

        # Round values
        for col in [f"{revtable}_LATEST", f"{revtable}_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Rename columns for clarity
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("HTOTALREV comparison completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing HTOTALREV: {e}")
        raise

def compare_htotalrev(transsheet_previous, transsheet_latest, ubersheet_previous, ubersheet_latest):
    try:
        logger.info("Comparing Hourly Total Revs between previous and latest sheets.")

        # Group by "PARTNER NAME" and sum "Total Rev"
        transprevious_grouped = transsheet_previous.groupby("PARTNER NAME", as_index=False)["TTL Rev"].sum()
        translatest_grouped = transsheet_latest.groupby("PARTNER NAME", as_index=False)["TTL Rev"].sum()
        
        uberprevious_grouped = ubersheet_previous.groupby("PARTNER NAME", as_index=False)["Total Rev"].sum()
        uberlatest_grouped = ubersheet_latest.groupby("PARTNER NAME", as_index=False)["Total Rev"].sum()
        
        # Merge both grouped dataframes on "PARTNER NAME"
        transcomparison = translatest_grouped.merge(transprevious_grouped, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")).fillna(0)
        ubercomparison = uberlatest_grouped.merge(uberprevious_grouped, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")).fillna(0)
        
        # Combine trans and uber revenue comparisons
        comparison = transcomparison.merge(ubercomparison, on="PARTNER NAME", how="outer", suffixes=("_TRANS", "_UBER")).fillna(0)
        
        # Calculate the total revenue change
        comparison["LATEST"] = comparison["TTL Rev_LATEST"] + comparison["Total Rev_LATEST"]
        comparison["PREVIOUS"] = comparison["TTL Rev_PREVIOUS"] + comparison["Total Rev_PREVIOUS"]
        comparison["CHANGE"] = comparison["LATEST"] - comparison["PREVIOUS"]
        
        # Round values
        for col in ["LATEST", "PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)
        
        # Rename columns for clarity
        comparison = comparison[["PARTNER NAME", "LATEST", "PREVIOUS", "CHANGE"]]
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]
        
        logger.info("HTOTALREV comparison completed.")
        return comparison
    
    except Exception as e:
        logger.error(f"Error comparing HTOTALREV: {e}")
        raise

def compare_liftlease(sheet_previous, sheet_latest, htotalrev_df):
    try:
        logger.info("Comparing Lift Lease between previous and latest sheets.")
        previous_values = None
        latest_values = None
        # if client == "TRANSDEV":
        #     previous_values = sheet_previous.groupby("PARTNER", as_index=False)["LIFT LEASE TOTAL"].sum()
        #     latest_values = sheet_latest.groupby("PARTNER", as_index=False)["LIFT LEASE TOTAL"].sum()
        # else:
        #     previous_grouped = sheet_previous.groupby("PARTNER NAME", as_index=False)["Total Rev"].sum()
        #     latest_grouped = sheet_latest.groupby("PARTNER NAME", as_index=False)["Total Rev"].sum()
        
        # Group by PARTNER and sum the LIFT LEASE TOTAL
        previous_values = sheet_previous.groupby("PARTNER", as_index=False)["LIFT LEASE TOTAL"].sum()
        latest_values = sheet_latest.groupby("PARTNER", as_index=False)["LIFT LEASE TOTAL"].sum()


        # Merge both dataframes on "PARTNER" and handle missing values with 0
        comparison = latest_values.merge(previous_values, on="PARTNER", how="outer", suffixes=("_LATEST", "_PREVIOUS")).fillna(0)
        comparison = comparison.merge(htotalrev_df, on="PARTNER", how="inner").fillna(0)

        # Calculate the change in the "LIFT LEASE TOTAL"
        comparison["CHANGE"] = comparison["LIFT LEASE TOTAL_LATEST"] - comparison["LIFT LEASE TOTAL_PREVIOUS"]
        for col in ["LIFT LEASE TOTAL_LATEST", "LIFT LEASE TOTAL_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Prepare the final dataframe for comparison
        deductions_comparison = comparison[["PARTNER", "LIFT LEASE TOTAL_LATEST", "LIFT LEASE TOTAL_PREVIOUS", "CHANGE"]].drop_duplicates(subset="PARTNER")

        # Rename columns
        deductions_comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("Lift Lease comparison completed.")
        return deductions_comparison

    except Exception as e:
        logger.error(f"Error comparing Lift Lease: {e}")
        raise

def compare_violations(sheet_previous, sheet_latest, htotalrev_df):
    try:
        logger.info("Comparing Lift Lease between previous and latest sheets.")
        
        # Group by PARTNER and sum the Violation
        previous_values = sheet_previous.groupby("PARTNER", as_index=False)["Violation"].sum()
        latest_values = sheet_latest.groupby("PARTNER", as_index=False)["Violation"].sum()


        # Merge both dataframes on "PARTNER" and handle missing values with 0
        comparison = latest_values.merge(previous_values, on="PARTNER", how="outer", suffixes=("_LATEST", "_PREVIOUS")).fillna(0)
        comparison = comparison.merge(htotalrev_df, on="PARTNER", how="inner").fillna(0)

        # Calculate the change in the "LIFT LEASE TOTAL"
        comparison["CHANGE"] = comparison["Violation_LATEST"] - comparison["Violation_PREVIOUS"]

        for col in ["Violation_LATEST", "Violation_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Prepare the final dataframe for comparison
        deductions_comparison = comparison[["PARTNER", "Violation_LATEST", "Violation_PREVIOUS", "CHANGE"]].drop_duplicates(subset="PARTNER")

        # Rename columns
        deductions_comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("Violation comparison completed.")
        return deductions_comparison

    except Exception as e:
        logger.error(f"Error comparing Violation: {e}")
        raise

def compare_operators(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing operators between previous and latest sheets.")
        
        # Extract and store unique pairs of (OPERATOR NAME, PARTNER NAME)
        operators_previous = set(sheet_previous[["OPERATOR NAME", "PARTNER NAME"]].dropna().itertuples(index=False, name=None))
        operators_latest = set(sheet_latest[["OPERATOR NAME", "PARTNER NAME"]].dropna().itertuples(index=False, name=None))

        # Identify added and removed operators
        added = operators_latest - operators_previous
        removed = operators_previous - operators_latest

        # Convert to DataFrame format
        added_list = [{"Operator Name": op, "Partner": partner, "Change": "Added"} for op, partner in added]
        removed_list = [{"Operator Name": op, "Partner": partner, "Change": "Removed"} for op, partner in removed]

        # Create DataFrames
        added_df = pd.DataFrame(added_list)
        removed_df = pd.DataFrame(removed_list)

        # Ensure DataFrames always have the necessary columns
        if added_df.empty:
            added_df = pd.DataFrame(columns=["Operator Name", "Partner", "Change"])
        if removed_df.empty:
            removed_df = pd.DataFrame(columns=["Operator Name", "Partner", "Change"])

        # Combine the results
        operator_changes_df = pd.concat([added_df, removed_df], ignore_index=True)

        logger.info("Operator comparison completed.")
        return operator_changes_df

    except Exception as e:
        logger.error(f"Error comparing operators: {e}")
        raise

def compare_acceptance_rate(sheet_previous, sheet_latest, week):
    try:
        logger.info(f"Comparing Acceptance Rate for {week}.")

        # Filter data for the specific week
        prev_week = sheet_previous[sheet_previous["WeekN"] == week]
        latest_week = sheet_latest[sheet_latest["WeekN"] == week]

        # Group by "PARTNER NAME" and mean Acceptance Rate
        prev_values = prev_week.groupby("PARTNER NAME", as_index=False)["Acceptance Rate"].mean()
        latest_values = latest_week.groupby("PARTNER NAME", as_index=False)["Acceptance Rate"].mean()

        # Merge both datasets
        comparison = latest_values.merge(
            prev_values, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")
        ).fillna(0)

        # Calculate the change
        comparison["CHANGE"] = comparison["Acceptance Rate_LATEST"] - comparison["Acceptance Rate_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["Acceptance Rate_LATEST", "Acceptance Rate_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].apply(lambda x: float(f"{x * 100:.2f}"))

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info(f"{week} Acceptance Rate comparison completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Acceptance Rate for {week}: {e}")
        raise


def compare_cancellation_rate(sheet_previous, sheet_latest, week):
    try:
        logger.info(f"Comparing Cancellation Rate for {week}.")

        # Filter data for the specific week
        prev_week = sheet_previous[sheet_previous["WeekN"] == week]
        latest_week = sheet_latest[sheet_latest["WeekN"] == week]

        # Group by "PARTNER NAME" and sum Cancellation Rate
        prev_values = prev_week.groupby("PARTNER NAME", as_index=False)["Cancellation Rate"].mean()
        latest_values = latest_week.groupby("PARTNER NAME", as_index=False)["Cancellation Rate"].mean()

        # Merge both datasets
        comparison = latest_values.merge(
            prev_values, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")
        ).fillna(0)

        # Calculate the change
        comparison["CHANGE"] = comparison["Cancellation Rate_LATEST"] - comparison["Cancellation Rate_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["Cancellation Rate_LATEST", "Cancellation Rate_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].apply(lambda x: float(f"{x * 100:.2f}"))

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info(f"{week} Cancellation Rate comparison completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Cancellation Rate for {week}: {e}")
        raise


def compare_utilization(sheet_previous, sheet_latest, week):
    try:
        logger.info(f"Comparing Utilization% for {week}.")

        # Filter data for the specific week
        prev_week = sheet_previous[sheet_previous["WeekN"] == week]
        latest_week = sheet_latest[sheet_latest["WeekN"] == week]

        # Group by "PARTNER NAME" and sum Utilization%
        prev_values = prev_week.groupby("PARTNER NAME", as_index=False)["Utilization%"].mean()
        latest_values = latest_week.groupby("PARTNER NAME", as_index=False)["Utilization%"].mean()

        # Merge both datasets
        comparison = latest_values.merge(
            prev_values, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")
        ).fillna(0)

        # Calculate the change
        comparison["CHANGE"] = comparison["Utilization%_LATEST"] - comparison["Utilization%_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["Utilization%_LATEST", "Utilization%_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].apply(lambda x: float(f"{x * 100:.2f}"))

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info(f"{week} Utilization% comparison completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Utilization% for {week}: {e}")
        raise

def compare_ReqHours(sheet_previous, sheet_latest, week):
    try:
        logger.info(f"Comparing Required Hours for {week}.")

        # Filter data for the specific week
        prev_week = sheet_previous[sheet_previous["WeekN"] == week]
        latest_week = sheet_latest[sheet_latest["WeekN"] == week]

        # Group by "PARTNER NAME" and sum Payable Online Hours
        prev_values = prev_week.groupby("PARTNER NAME", as_index=False)["% of Hours to Required"].mean()
        latest_values = latest_week.groupby("PARTNER NAME", as_index=False)["% of Hours to Required"].mean()

        # Merge both datasets
        comparison = latest_values.merge(
            prev_values, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")
        ).fillna(0)

        # Calculate the change
        comparison["CHANGE"] = comparison["% of Hours to Required_LATEST"] - comparison["% of Hours to Required_PREVIOUS"]
        for col in ["% of Hours to Required_LATEST", "% of Hours to Required_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].apply(lambda x: float(f"{x * 100:.2f}"))

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info(f"{week} Required Hours comparison completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Required Hours for {week}: {e}")
        raise

def compare_pOnlineHours(sheet_previous, sheet_latest, week):
    try:
        logger.info(f"Comparing Payable Online Hours for {week}.")

        # Filter data for the specific week
        prev_week = sheet_previous[sheet_previous["WeekN"] == week]
        latest_week = sheet_latest[sheet_latest["WeekN"] == week]

        # Group by "PARTNER NAME" and sum Payable Online Hours
        prev_values = prev_week.groupby("PARTNER NAME", as_index=False)["Payable Online Hours"].sum()
        latest_values = latest_week.groupby("PARTNER NAME", as_index=False)["Payable Online Hours"].sum()

        # Merge both datasets
        comparison = latest_values.merge(
            prev_values, on="PARTNER NAME", how="outer", suffixes=("_LATEST", "_PREVIOUS")
        ).fillna(0)

        # Calculate the change
        comparison["CHANGE"] = comparison["Payable Online Hours_LATEST"] - comparison["Payable Online Hours_PREVIOUS"]
        for col in ["Payable Online Hours_LATEST", "Payable Online Hours_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].apply(lambda x: float(f"{x:.2f}"))

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info(f"{week} Payable Online Hours comparison completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Payable Online Hours for {week}: {e}")
        raise

def find_missing_dates(sheet_previous, sheet_latest):
    try:
        logger.info("Checking for missing dates in the consecutive range.")
        
        # Extracting the Date column and dropping NaN values
        dates_previous = set(pd.to_datetime(sheet_previous["Date"].dropna()))
        dates_latest = set(pd.to_datetime(sheet_latest["Date"].dropna()))
        
        # Finding the full expected date range
        all_dates = pd.date_range(min(dates_previous.union(dates_latest)), 
                                  max(dates_previous.union(dates_latest)))
        
        # Finding missing dates
        missing_dates = sorted(set(all_dates) - dates_previous - dates_latest)
        
        # Converting to DataFrame
        if not missing_dates:
            missing_df = pd.DataFrame({"Missing Dates of Work": ["There are no missing dates of work."]})
        else:
            # missing_df = pd.DataFrame(missing_dates, columns=["Missing Dates"])
            missing_df = pd.DataFrame([date.strftime("-----  %B %d, %Y  -----") for date in missing_dates], columns=["Missing Dates of Work"])
        
        logger.info("Missing date check completed.")
        return missing_df
    except Exception as e:
        logger.error(f"Error finding missing dates: {e}")
        raise


def compare_trips(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing trips data between previous and latest sheets.")

        # Group by "PARTNER NAME" and mean Acceptance Rate
        prev_values = sheet_previous.groupby("Date", as_index=False)["Trips"].sum()
        latest_values = sheet_latest.groupby("Date", as_index=False)["Trips"].sum()

        # Merge both datasets
        comparison = latest_values.merge(
            prev_values, on="Date", how="outer", suffixes=("_LATEST", "_PREVIOUS")
        ).fillna(0)

        # Calculate the change
        comparison["CHANGE"] = comparison["Trips_LATEST"] - comparison["Trips_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["Trips_LATEST", "Trips_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Rename columns
        comparison.columns = ["DATE", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("Trips comparison completed.")
        return comparison
    except Exception as e:
        logger.error(f"Error comparing trips: {e}")
        raise

def apply_formatting(sheet_name, wb):
    try:
        logger.info(f"Applying formatting to sheet: {sheet_name}.")
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                if ws[1][cell.column - 1].value.lower() == "change":
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100")
                        elif cell.value < 0:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006")
                    elif isinstance(cell.value, str):
                        if cell.value.lower() in ["increased", "added"]:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100")
                        elif cell.value.lower() in ["decreased", "removed"]:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006")
        logger.info(f"Formatting applied successfully to sheet: {sheet_name}.")
    except Exception as e:
        logger.error(f"Error applying formatting to sheet {sheet_name}: {e}")
        raise


def save_comparison_results(output_folder, comparison_data, filename):
    try:
        logger.info(f"Saving comparison results to {filename}.")
        os.makedirs(output_folder, exist_ok=True)
        full_comparison_file = os.path.join(output_folder, filename)
        with pd.ExcelWriter(full_comparison_file, engine="openpyxl") as writer:
            for sheet_name, data in comparison_data.items():
                data.to_excel(writer, sheet_name=sheet_name, index=False)

        wb_full = load_workbook(full_comparison_file)
        for sheet in comparison_data.keys():
            apply_formatting(sheet, wb_full)
        wb_full.save(full_comparison_file)
        wb_full.close()

        logger.info(f"Comparison results saved successfully to {filename}.")
    except Exception as e:
        logger.error(f"Error saving comparison results: {e}")
        raise

def main(file_previous, file_latest):
    try: 
        file_entry_previous = file_previous
        file_entry_latest = file_latest

        print(f"{file_entry_previous} + {file_entry_latest}")
        logger.info("Starting main comparison process.")
        output_folder = "ComparedResults"
        os.makedirs(output_folder, exist_ok=True)

        sheet_pr_previous, sheet_pr_latest, sheet_vdpmv_previous, sheet_vdpmv_latest, sheet_trapeze_previous, sheet_trapeze_latest, sheet_deductions_previous, sheet_deductions_latest = load_sheets(file_previous, file_latest)

        unique_clients = sheet_deductions_latest["Type"].dropna().unique()

        for client in unique_clients:
            # 1. Process the data without any filtering (full comparison)
            sheet_cdeductions_latest = sheet_deductions_latest[sheet_deductions_latest["Type"] == client]
            sheet_cdeductions_previous = sheet_deductions_previous[sheet_deductions_previous["Type"] == client]
            # logger.info(f"Sheet Deductions Latest: {sheet_cdeductions_latest}")
            # logger.info(f"Sheet Deductions Previous: {sheet_cdeductions_previous}")
            # Compare totals

            t_client = ""
            htotalprev = None
            htotallatest = None
            logger.info(f"Total Clients: {client}")
            # 2. Compare hourly total revs
            if client == "TRANSDEV":
                t_client = "DIV10_TRIMET"
                htotalprev = sheet_trapeze_previous
                htotallatest = sheet_trapeze_latest
                logger.info(f"Client in total: {t_client}")
            elif client == "UBER":
                htotalprev = sheet_vdpmv_previous
                htotallatest = sheet_vdpmv_latest
                t_client = "UBER WAV TRANSIT"
                logger.info(f"Client in total: {t_client}")

            prev_totals = calculate_client_totals(sheet_cdeductions_previous, sheet_pr_previous, t_client)
            lat_totals = calculate_client_totals(sheet_cdeductions_latest, sheet_pr_latest, t_client)
            totals_comparison_df = compare_totals(prev_totals, lat_totals)
            
            compare_htotalrev_df = compare_client_htotalrev(htotalprev, htotallatest, client)
            # logger.info(f"HTOTAL_REV DF: {compare_htotalrev_df}")

            # 3. Compare Lift Lease
            compare_liftlease_df = compare_liftlease(sheet_cdeductions_previous, sheet_cdeductions_latest, compare_htotalrev_df)
            # logger.info(f"Lift Lease DF: {compare_liftlease_df}")

            # 4. Compare Violations
            compare_violations_df = compare_violations(sheet_cdeductions_previous, sheet_cdeductions_latest, compare_htotalrev_df)
            # logger.info(f"Violations DF: {compare_violations_df}")

            # 5. Compare Trips
            compare_trips_df = compare_trips(sheet_trapeze_previous, sheet_trapeze_latest)
            # logger.info(f"Trips DF: {compare_trips_df}")

            # 5. Compare operators
            # operator_changes_df = compare_operators(sheet_partner_previous, sheet_partner_latest)
            # logger.info(f"Operator Changes DF:{operator_changes_df}")


            # 6. Week 1 comparison
            # Save the full comparison results
            full_comparison_file = os.path.join(output_folder, f"DIV10_{client}_Tables.xlsx")
            excel_sheets = []
            with pd.ExcelWriter(full_comparison_file, engine="openpyxl") as writer:
                totals_comparison_df.to_excel(writer, sheet_name="TotalInvoicePayment", index=False)
                excel_sheets.append("TotalInvoicePayment")
                compare_htotalrev_df.to_excel(writer, sheet_name="HTotalRevComparison", index=False)
                excel_sheets.append("HTotalRevComparison")
                compare_liftlease_df.to_excel(writer, sheet_name="LiftLeaseComparison", index=False)
                excel_sheets.append("LiftLeaseComparison")
                compare_violations_df.to_excel(writer, sheet_name="ViolationComparison", index=False)
                excel_sheets.append("ViolationComparison")
                if client == "TRANSDEV":
                    compare_trips_df.to_excel(writer, sheet_name="TripsComparison", index=False)
                    excel_sheets.append("TripsComparison")
                # operator_changes_df.to_excel(writer, sheet_name="OperatorChanges", index=False)
                # excel_sheets.append("OperatorChanges")

            # Apply formatting to the full comparison file
            wb_full = load_workbook(full_comparison_file)
            for sheet in excel_sheets:                                                                                                
                apply_formatting(sheet, wb_full)
            wb_full.save(full_comparison_file)
        wb_full.close()
        # return 0
                

# 1. Process the data without any filtering (full comparison)

        # Compare totals
        prev_totals = calculate_totals(sheet_deductions_previous, sheet_pr_previous)
        lat_totals = calculate_totals(sheet_deductions_latest, sheet_pr_latest)
        totals_comparison_df = compare_totals(prev_totals, lat_totals)
        # logger.info(f"Totals DF: {totals_comparison_df}")

        # 2. Compare hourly total revs
        compare_htotalrev_df = compare_htotalrev(sheet_trapeze_previous, sheet_trapeze_latest, sheet_vdpmv_previous, sheet_vdpmv_latest)
        # logger.info(f"HTOTAL_REV DF: {compare_htotalrev_df}")

        # 3. Compare Lift Lease
        compare_liftlease_df = compare_liftlease(sheet_deductions_previous, sheet_deductions_latest, compare_htotalrev_df)
        # logger.info(f"Lift Lease DF: {compare_liftlease_df}")

        # 4. Compare Violations
        compare_violations_df = compare_violations(sheet_deductions_previous, sheet_deductions_latest, compare_htotalrev_df)
        # logger.info(f"Violations DF: {compare_violations_df}")

        # 5. Compare operators
        # operator_changes_df = compare_operators(sheet_partner_previous, sheet_partner_latest)
        # logger.info(f"Operator Changes DF:{operator_changes_df}")

        # 6. Week 1 comparison
            

        # Save the full comparison results
        full_comparison_file = os.path.join(output_folder, "DIV10_Tables.xlsx")
        excel_sheets = []
        with pd.ExcelWriter(full_comparison_file, engine="openpyxl") as writer:
            totals_comparison_df.to_excel(writer, sheet_name="TotalInvoicePayment", index=False)
            excel_sheets.append("TotalInvoicePayment")
            compare_htotalrev_df.to_excel(writer, sheet_name="HTotalRevComparison", index=False)
            excel_sheets.append("HTotalRevComparison")
            compare_liftlease_df.to_excel(writer, sheet_name="LiftLeaseComparison", index=False)
            excel_sheets.append("LiftLeaseComparison")
            compare_violations_df.to_excel(writer, sheet_name="ViolationComparison", index=False)
            excel_sheets.append("ViolationComparison")
            # operator_changes_df.to_excel(writer, sheet_name="OperatorChanges", index=False)
            # excel_sheets.append("OperatorChanges")
            for week in ["Week 1", "Week 2"]:
                acceptance_changes_df = compare_acceptance_rate(sheet_vdpmv_previous, sheet_vdpmv_latest, week)
                acceptance_changes_df.to_excel(writer, sheet_name=f"{week}AcceptRateComp", index=False)
                excel_sheets.append(f"{week}AcceptRateComp")

                cancellation_changes_df = compare_cancellation_rate(sheet_vdpmv_previous, sheet_vdpmv_latest, week)
                cancellation_changes_df.to_excel(writer, sheet_name=f"{week}CancelRateComp", index=False)
                excel_sheets.append(f"{week}CancelRateComp")

                utilization_changes_df = compare_utilization(sheet_vdpmv_previous, sheet_vdpmv_latest, week)
                utilization_changes_df.to_excel(writer, sheet_name=f"{week}UtilizationComp", index=False)
                excel_sheets.append(f"{week}UtilizationComp")

                pOnlinehours_changes_df = compare_pOnlineHours(sheet_vdpmv_previous, sheet_vdpmv_latest, week)
                pOnlinehours_changes_df.to_excel(writer, sheet_name=f"{week}POnlineHrsComp", index=False)
                excel_sheets.append(f"{week}POnlineHrsComp")

                # pbonushours_changes_df = compare_pBonusHours(sheet_vdpmv_previous, sheet_vdpmv_latest, week)
                # pbonushours_changes_df.to_excel(writer, sheet_name=f"{week}PBonusHrsComp", index=False)
                # excel_sheets.append(f"{week}PBonusHrsComp")

                reqhours_changes_df = compare_ReqHours(sheet_vdpmv_previous, sheet_vdpmv_latest, week)
                reqhours_changes_df.to_excel(writer, sheet_name=f"{week}ReqHrsComp", index=False)
                excel_sheets.append(f"{week}ReqHrsComp")

        # Doperator_changes_df.to_excel(writer, sheet_name="DateComparison", index=False)

        # Apply formatting to the full comparison file
        wb_full = load_workbook(full_comparison_file)
        for sheet in excel_sheets:                                                                                                
            apply_formatting(sheet, wb_full)
        wb_full.save(full_comparison_file)
        wb_full.close()

        logger.info(f"Main comparison process completed successfully. File saved to {full_comparison_file}.")
        time.sleep(2)
        db.main(file_previous, file_latest)
    except Exception as e:
        logger.error(f"Error in main comparison process: {e}")
        raise
    # finally:
    #     wb_client.close()


def open_file_dialog(entry):
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsm;*.xlsx")])
    if filename:
        entry.delete(0, tk.END)
        entry.insert(0, filename)


def create_gui():
    global file_entry_previous, file_entry_latest # Declare them as global variables

    # Set up the GUI window
    root = tk.Tk()
    root.title("Comparison Tool")

    # Create and place labels, entry boxes, and buttons
    tk.Label(root, text="Previous File:").grid(row=0, column=0, padx=10, pady=5)
    entry_previous = tk.Entry(root, width=50)
    entry_previous.grid(row=0, column=1, padx=10, pady=5)
    
    # tk.Button(root, text="Browse", command=lambda: open_file_dialog(entry_previous)).grid(row=0, column=2, padx=10, pady=5)

    tk.Label(root, text="Latest File:").grid(row=1, column=0, padx=10, pady=5)
    entry_latest = tk.Entry(root, width=50)
    entry_latest.grid(row=1, column=1, padx=10, pady=5)
    
    # tk.Button(root, text="Browse", command=lambda: open_file_dialog(entry_latest)).grid(row=1, column=2, padx=10, pady=5)

    # Button to trigger the comparison process
    # tk.Button(root, text="Compare", command=lambda: (main(entry_previous.get(), entry_latest.get()), root.destroy())).grid(row=2, column=1, pady=20)
    tk.Button(root, text="Compare", command=lambda: handle_comparison(entry_previous.get(), entry_latest.get(), root)).grid(row=2, column=1, pady=20)

    def handle_comparison(file_previous, file_latest, root):
        try:
            main(file_previous, file_latest)
        except Exception as e:
            print(f"An error occurred: {e}")
            # Check for the disconnection error and close the GUI if it happens
            if isinstance(e, OSError) and "The object invoked has disconnected" in str(e):
                print("Disconnected from Excel, closing GUI.")
                root.quit()  # This will close the Tkinter window
        finally:
            root.destroy()  # Close the window in all cases
    # Start the GUI loop
    root.mainloop()


if __name__ == "__main__":
    create_gui()

